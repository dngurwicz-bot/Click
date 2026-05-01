import base64
import csv
import html as html_module
import io
import json
import uuid
from collections import defaultdict
from datetime import UTC, date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.exc import ProgrammingError
from sqlalchemy.ext.asyncio import AsyncSession

from app.middleware.auth import CurrentUser
from app.models.admin_user import AdminUser
from app.models.admin_user_permission import AdminUserPermission
from app.models.audit_log import AuditLog
from app.models.module import Module, ModulePrice, OrgTemplate, OrgTemplateDefault, OrgTemplateModule
from app.models.saved_report_view import SavedReportView
from app.models.tenant import (
    Tenant,
    TenantAddress,
    TenantContact,
    TenantIdentity,
    TenantStatus,
    TenantSubscription,
    TenantSubscriptionModule,
)
from app.schemas.reporting import (
    ReportCatalogItem,
    ReportCatalogResponse,
    ReportDatasetDefinition,
    ReportDatasetsResponse,
    ReportDefinition,
    ReportExportRequest,
    ReportExportResponse,
    ReportFieldDefinition,
    ReportFieldHelp,
    ReportFilterOptions,
    ReportFilterOption,
    ReportMetricDefinition,
    ReportMetricRequest,
    ReportMetricValue,
    ReportQueryRequest,
    ReportResult,
    SavedReportViewCreate,
    SavedReportViewOut,
    SavedReportViewUpdate,
)
from app.services.subscription_modules import derive_subscription_snapshot

FONT_NAME = "NotoSansHebrew"
FONT_PATH = Path(__file__).resolve().parent.parent / "assets" / "fonts" / "NotoSansHebrew-Regular.ttf"
TENANT_STATUS_OPTIONS = [
    ReportFilterOption(value="active", label="פעיל"),
    ReportFilterOption(value="trial", label="ניסיון"),
    ReportFilterOption(value="suspended", label="מושהה"),
    ReportFilterOption(value="cancelled", label="מבוטל"),
]
DEFAULT_OPERATORS_BY_TYPE = {
    "string": ["equals", "not_equals", "contains", "is_null", "is_not_null", "in", "not_in"],
    "number": ["equals", "not_equals", "greater_than", "greater_or_equal", "less_than", "less_or_equal", "is_null", "is_not_null"],
    "date": ["equals", "greater_than", "greater_or_equal", "less_than", "less_or_equal", "is_null", "is_not_null"],
    "datetime": ["equals", "greater_than", "greater_or_equal", "less_than", "less_or_equal", "is_null", "is_not_null"],
    "uuid": ["equals", "not_equals", "is_null", "is_not_null", "in", "not_in"],
    "boolean": ["equals", "not_equals", "is_null", "is_not_null"],
}
OPERATOR_LABELS = {
    "equals": "שווה",
    "not_equals": "שונה",
    "contains": "מכיל",
    "greater_than": "גדול מ",
    "greater_or_equal": "גדול או שווה",
    "less_than": "קטן מ",
    "less_or_equal": "קטן או שווה",
    "is_null": "ריק",
    "is_not_null": "לא ריק",
    "in": "ברשימה",
    "not_in": "לא ברשימה",
}
LEGACY_DATASET_ALIASES = {
    "tenant_snapshot": "tenant_snapshot_full",
    "tenant_module_snapshot": "tenant_module_snapshot_full",
}


def _register_font() -> None:
    try:
        pdfmetrics.getFont(FONT_NAME)
    except KeyError:
        pdfmetrics.registerFont(TTFont(FONT_NAME, str(FONT_PATH)))


def _rtl(text: str | None) -> str:
    value = str(text or "")
    return get_display(value) if any("\u0590" <= ch <= "\u05FF" for ch in value) else value


def _fmt_decimal(value: Decimal | int | float | None) -> str:
    if value is None:
        return "0"
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    return f"{value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):,}"


def _serialize_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _fmt_value(value: Any) -> str:
    value = _serialize_value(value)
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "כן" if value else "לא"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return _fmt_decimal(value)
    return str(value)


def _seat_bucket(seat_count: int) -> str:
    if seat_count <= 10:
        return "0-10"
    if seat_count <= 25:
        return "11-25"
    if seat_count <= 50:
        return "26-50"
    if seat_count <= 100:
        return "51-100"
    return "101+"


def _metric_label(metric: ReportMetricRequest) -> str:
    if metric.label:
        return metric.label
    if metric.operation == "count":
        return "כמות רשומות"
    if metric.operation == "count_distinct":
        return f"כמות ייחודית: {metric.field or 'value'}"
    return f"{metric.operation.upper()} {metric.field or ''}".strip()


def _field(
    field_id: str,
    label: str,
    field_type: str,
    *,
    category: str,
    description: str | None = None,
    operators: list[str] | None = None,
    groupable: bool | None = None,
) -> ReportFieldDefinition:
    return ReportFieldDefinition(
        id=field_id,
        label=label,
        type=field_type,
        operators=operators or DEFAULT_OPERATORS_BY_TYPE[field_type],
        groupable=(field_type in {"string", "date", "datetime", "boolean"} if groupable is None else groupable),
        category=category,
        description=description,
        help=ReportFieldHelp(summary=description or label, details=description or label),
    )


FIELD_GROUP_PREFIXES = [
    "identity_",
    "contact_main_",
    "address_main_",
    "status_",
    "subscription_",
    "module_",
    "template_",
    "saved_report_",
    "audit_",
    "admin_",
    "permission_",
    "period_",
]


def _field_group_key(field_id: str) -> str:
    for prefix in FIELD_GROUP_PREFIXES:
        if field_id.startswith(prefix):
            return prefix
    if field_id in {"tenant_id", "tenant_name", "tenant_status", "org_number"}:
        return "tenant_core"
    if field_id in {"valid_from", "valid_to", "created_at", "created_by"}:
        return "generic_timeline"
    return field_id


def _generate_field_summary(field: ReportFieldDefinition) -> str:
    if field.description:
        return field.description
    if field.id.endswith("_id"):
        return f"מזהה טכני של {field.label} לצורכי קישור, מעקב וסינון."
    if field.id.endswith("_name"):
        return f"שם התצוגה של {field.label}, כפי שמופיע בדוח ובמסכי הניהול."
    if field.id.endswith("_slug"):
        return f"מזהה יציב של {field.label} שמשמש לקישור בין מודולים, תמחור ותהליכים אוטומטיים."
    if field.id.endswith("_created_at"):
        return f"זמן היצירה של {field.label} לצורכי היסטוריה, מיון ובקרה."
    if field.id.endswith("_updated_at"):
        return f"זמן העדכון האחרון של {field.label} לצורכי מעקב היסטורי."
    if field.id.endswith("_valid_from"):
        return f"תאריך תחילת התוקף של {field.label} בתוך ציר הזמן של הרשומה."
    if field.id.endswith("_valid_to"):
        return f"תאריך סיום התוקף של {field.label}; ערך ריק בדרך כלל אומר שהרשומה עדיין פעילה."
    if field.type == "boolean":
        return f"שדה כן/לא שמציין את המצב של {field.label}."
    if field.type == "number":
        return f"שדה מספרי המשמש למדידה, חישוב וסיכום עבור {field.label}."
    if field.type in {"date", "datetime"}:
        return f"שדה תאריך/זמן המשמש למעקב, סינון ומיון עבור {field.label}."
    return f"שדה מידע מסוג {field.type} עבור {field.label}."


def _find_related_fields(field: ReportFieldDefinition, fields: list[ReportFieldDefinition]) -> tuple[list[str], str | None]:
    field_ids = {item.id for item in fields}
    related: list[str] = []
    group_key = _field_group_key(field.id)

    def add(candidate: str) -> None:
        if candidate in field_ids and candidate != field.id and candidate not in related:
            related.append(candidate)

    if group_key not in {"tenant_core", "generic_timeline", field.id}:
        for item in fields:
            if _field_group_key(item.id) == group_key and item.id != field.id:
                related.append(item.id)
            if len(related) >= 4:
                break

    pairs = [
        ("_valid_from", "_valid_to"),
        ("_created_by", "_created_at"),
        ("_updated_by", "_updated_at"),
    ]
    for left, right in pairs:
        if field.id.endswith(left):
            add(field.id[: -len(left)] + right)
        if field.id.endswith(right):
            add(field.id[: -len(right)] + left)

    if field.id.endswith("_id"):
        base = field.id[:-3]
        add(base + "_name")
        add(base + "_slug")
        add(base + "_dataset")

    if field.id.endswith("_slug"):
        base = field.id[:-5]
        add(base + "_name")
        add(base + "_description")

    explicit_relations = {
        "org_number": ["tenant_id", "identity_name_he"],
        "tenant_id": ["org_number", "identity_name_he", "tenant_name"],
        "tenant_name": ["tenant_id", "tenant_status"],
        "tenant_status": ["tenant_name", "subscription_id"],
        "subscription_template_id": ["subscription_template_name"],
        "subscription_template_name": ["subscription_template_id"],
        "module_count": ["module_names", "subscription_selected_module_slugs"],
        "module_names": ["module_count", "subscription_selected_module_slugs"],
        "module_seats": ["subscription_seat_count", "module_name"],
        "pricing_mode": ["override_base_price_ils", "override_per_seat_ils", "base_price_ils", "per_seat_ils"],
        "base_price_ils": ["per_seat_ils", "included_seats", "pricing_mode"],
        "override_base_price_ils": ["override_per_seat_ils", "override_included_seats", "price_lock_reason"],
        "template_name": ["template_id"],
        "default_value": ["default_type", "is_mandatory"],
        "tenant_count": ["total_seats", "avg_seats"],
        "total_seats": ["tenant_count", "avg_seats"],
        "avg_seats": ["tenant_count", "total_seats"],
    }
    for candidate in explicit_relations.get(field.id, []):
        add(candidate)

    related_reason = None
    if related:
        related_reason = "השדות המקושרים עוזרים להבין את ההקשר העסקי, מקור הרשומה או ציר הזמן של אותו אובייקט."
    return related[:4], related_reason


def _build_field_help(field: ReportFieldDefinition, fields: list[ReportFieldDefinition]) -> ReportFieldHelp:
    related_fields, related_reason = _find_related_fields(field, fields)
    operator_labels = [OPERATOR_LABELS.get(operator, operator) for operator in field.operators]
    notes = [
        f"סוג נתון: {field.type}.",
        f"ניתן לסנן באמצעות: {', '.join(operator_labels)}.",
        "ניתן להשתמש בשדה הזה בקיבוץ וסיכומים." if field.groupable else "השדה לא מיועד לקיבוץ, ורצוי להשתמש בו בעיקר כתצוגה או סינון.",
    ]
    if field.id.endswith("_valid_to"):
        notes.append("ערך ריק בדרך כלל מייצג רשומה פתוחה או פעילה ללא תאריך סיום.")
    if field.id.endswith("_id"):
        notes.append("זהו בדרך כלל מזהה מערכת פנימי, ולכן נוח לשילוב עם שדות שם/slug לצורך קריאה אנושית.")
    if field.type == "boolean":
        notes.append("בערכים בוליאניים התצוגה בממשק היא כן/לא, אך מאחורי הקלעים נשמר ערך true/false.")

    details = (
        f"השדה שייך לקטגוריית '{field.category or 'כללי'}' ונמצא במבנה הדוחות כדי לאפשר חיפוש, סינון, "
        f"מיון וייצוא ברמת {field.label}. "
        f"ברוב המקרים משתמשים בו יחד עם שדות מאותה משפחה כדי להבין גם את הזהות, גם את הסטטוס וגם את חלון התוקף של הרשומה."
    )
    return ReportFieldHelp(
        summary=_generate_field_summary(field),
        details=details,
        related_fields=related_fields,
        related_reason=related_reason,
        notes=notes,
    )


def _enrich_dataset_fields(fields: list[ReportFieldDefinition]) -> list[ReportFieldDefinition]:
    return [field.model_copy(update={"help": _build_field_help(field, fields)}) for field in fields]


TENANT_SNAPSHOT_FIELDS = [
    _field("tenant_id", "מזהה לקוח", "uuid", category="לקוח"),
    _field("org_number", "מספר ארגון", "number", category="לקוח", groupable=True),
    _field("tenant_created_at", "נוצר בתאריך", "datetime", category="לקוח"),
    _field("identity_id", "מזהה רשומת זהות", "uuid", category="זהות"),
    _field("identity_name_he", "שם לקוח", "string", category="זהות", groupable=True),
    _field("identity_name_en", "שם באנגלית", "string", category="זהות"),
    _field("identity_tax_id", "ח.פ / ע.מ", "string", category="זהות"),
    _field("identity_entity_type", "סוג ישות", "string", category="זהות", groupable=True),
    _field("identity_logo_url", "לוגו", "string", category="זהות"),
    _field("identity_industry_code", "ענף", "string", category="זהות", groupable=True),
    _field("identity_valid_from", "זהות מתאריך", "date", category="זהות"),
    _field("identity_valid_to", "זהות עד תאריך", "date", category="זהות"),
    _field("identity_created_by", "זהות נוצרה על ידי", "uuid", category="זהות"),
    _field("identity_created_at", "זהות נוצרה בתאריך", "datetime", category="זהות"),
    _field("identity_updated_by", "זהות עודכנה על ידי", "uuid", category="זהות"),
    _field("identity_updated_at", "זהות עודכנה בתאריך", "datetime", category="זהות"),
    _field("contact_main_id", "מזהה איש קשר", "uuid", category="איש קשר"),
    _field("contact_main_type", "סוג איש קשר", "string", category="איש קשר", groupable=True),
    _field("contact_main_name", "איש קשר ראשי", "string", category="איש קשר"),
    _field("contact_main_email", "דוא\"ל ראשי", "string", category="איש קשר"),
    _field("contact_main_phone", "טלפון ראשי", "string", category="איש קשר"),
    _field("contact_main_phone_alt", "טלפון נוסף", "string", category="איש קשר"),
    _field("contact_main_website", "אתר", "string", category="איש קשר"),
    _field("contact_main_valid_from", "איש קשר מתאריך", "date", category="איש קשר"),
    _field("contact_main_valid_to", "איש קשר עד תאריך", "date", category="איש קשר"),
    _field("contact_main_created_by", "איש קשר נוצר על ידי", "uuid", category="איש קשר"),
    _field("contact_main_created_at", "איש קשר נוצר בתאריך", "datetime", category="איש קשר"),
    _field("contact_main_updated_by", "איש קשר עודכן על ידי", "uuid", category="איש קשר"),
    _field("contact_main_updated_at", "איש קשר עודכן בתאריך", "datetime", category="איש קשר"),
    _field("address_main_id", "מזהה כתובת", "uuid", category="כתובת"),
    _field("address_main_type", "סוג כתובת", "string", category="כתובת", groupable=True),
    _field("address_main_street", "רחוב", "string", category="כתובת"),
    _field("address_main_city", "עיר", "string", category="כתובת", groupable=True),
    _field("address_main_zip_code", "מיקוד", "string", category="כתובת"),
    _field("address_main_country", "מדינה", "string", category="כתובת", groupable=True),
    _field("address_main_valid_from", "כתובת מתאריך", "date", category="כתובת"),
    _field("address_main_valid_to", "כתובת עד תאריך", "date", category="כתובת"),
    _field("address_main_created_by", "כתובת נוצרה על ידי", "uuid", category="כתובת"),
    _field("address_main_created_at", "כתובת נוצרה בתאריך", "datetime", category="כתובת"),
    _field("address_main_updated_by", "כתובת עודכנה על ידי", "uuid", category="כתובת"),
    _field("address_main_updated_at", "כתובת עודכנה בתאריך", "datetime", category="כתובת"),
    _field("status_id", "מזהה סטטוס", "uuid", category="סטטוס"),
    _field("status_value", "סטטוס לקוח", "string", category="סטטוס", groupable=True),
    _field("status_reason", "סיבת סטטוס", "string", category="סטטוס", groupable=True),
    _field("status_notes", "הערות סטטוס", "string", category="סטטוס"),
    _field("status_valid_from", "סטטוס מתאריך", "date", category="סטטוס"),
    _field("status_valid_to", "סטטוס עד תאריך", "date", category="סטטוס"),
    _field("status_created_by", "סטטוס נוצר על ידי", "uuid", category="סטטוס"),
    _field("status_created_at", "סטטוס נוצר בתאריך", "datetime", category="סטטוס"),
    _field("status_updated_by", "סטטוס עודכן על ידי", "uuid", category="סטטוס"),
    _field("status_updated_at", "סטטוס עודכן בתאריך", "datetime", category="סטטוס"),
    _field("subscription_id", "מזהה מנוי", "uuid", category="מנוי"),
    _field("subscription_billing_cycle", "מחזור חיוב", "string", category="מנוי", groupable=True),
    _field("subscription_currency", "מטבע", "string", category="מנוי", groupable=True),
    _field("subscription_template_id", "מזהה תבנית", "uuid", category="מנוי"),
    _field("subscription_template_name", "שם תבנית", "string", category="מנוי", groupable=True),
    _field("subscription_seat_count", "מושבים במנוי", "number", category="מנוי"),
    _field("subscription_selected_module_slugs", "מודולים נבחרים", "string", category="מנוי"),
    _field("subscription_discount_pct", "אחוז הנחה", "number", category="מנוי"),
    _field("subscription_is_price_locked", "מחיר נעול", "boolean", category="מנוי", groupable=True),
    _field("subscription_next_renewal_at", "חידוש הבא", "date", category="מנוי", groupable=True),
    _field("subscription_valid_from", "מנוי מתאריך", "date", category="מנוי"),
    _field("subscription_valid_to", "מנוי עד תאריך", "date", category="מנוי"),
    _field("subscription_created_by", "מנוי נוצר על ידי", "uuid", category="מנוי"),
    _field("subscription_created_at", "מנוי נוצר בתאריך", "datetime", category="מנוי"),
    _field("subscription_updated_by", "מנוי עודכן על ידי", "uuid", category="מנוי"),
    _field("subscription_updated_at", "מנוי עודכן בתאריך", "datetime", category="מנוי"),
    _field("module_count", "כמות מודולים", "number", category="מנוי"),
    _field("module_names", "שמות מודולים", "string", category="מנוי"),
]

TENANT_MODULE_FIELDS = [
    _field("tenant_id", "מזהה לקוח", "uuid", category="לקוח"),
    _field("org_number", "מספר ארגון", "number", category="לקוח", groupable=True),
    _field("tenant_name", "שם לקוח", "string", category="לקוח", groupable=True),
    _field("tenant_status", "סטטוס לקוח", "string", category="לקוח", groupable=True),
    _field("subscription_id", "מזהה מנוי", "uuid", category="מנוי"),
    _field("subscription_template_name", "שם תבנית", "string", category="מנוי", groupable=True),
    _field("subscription_billing_cycle", "מחזור חיוב", "string", category="מנוי", groupable=True),
    _field("subscription_currency", "מטבע", "string", category="מנוי", groupable=True),
    _field("subscription_seat_count", "מושבים במנוי", "number", category="מנוי"),
    _field("subscription_created_by", "מנוי נוצר על ידי", "uuid", category="מנוי"),
    _field("subscription_created_at", "מנוי נוצר בתאריך", "datetime", category="מנוי"),
    _field("subscription_updated_by", "מנוי עודכן על ידי", "uuid", category="מנוי"),
    _field("subscription_updated_at", "מנוי עודכן בתאריך", "datetime", category="מנוי"),
    _field("module_assignment_id", "מזהה שיוך מודול", "uuid", category="שיוך מודול"),
    _field("tenant_subscription_id", "מזהה מנוי בשיוך", "uuid", category="שיוך מודול"),
    _field("module_slug", "קוד מודול", "string", category="מודול", groupable=True),
    _field("module_name", "שם מודול", "string", category="מודול", groupable=True),
    _field("module_description", "תיאור מודול", "string", category="מודול"),
    _field("module_icon", "אייקון מודול", "string", category="מודול"),
    _field("module_color_hex", "צבע מודול", "string", category="מודול", groupable=True),
    _field("module_is_required", "מודול חובה", "boolean", category="מודול", groupable=True),
    _field("module_is_active", "מודול פעיל", "boolean", category="מודול", groupable=True),
    _field("module_sort_order", "סדר מודול", "number", category="מודול"),
    _field("module_depends_on", "תלות מודול", "string", category="מודול"),
    _field("source_type", "מקור שיוך", "string", category="שיוך מודול", groupable=True),
    _field("module_status", "סטטוס שיוך", "string", category="שיוך מודול", groupable=True),
    _field("module_seats", "מושבי מודול", "number", category="שיוך מודול"),
    _field("pricing_mode", "שיטת תמחור", "string", category="שיוך מודול", groupable=True),
    _field("base_price_ils", "מחיר בסיס קטלוגי", "number", category="תמחור"),
    _field("per_seat_ils", "מחיר למושב קטלוגי", "number", category="תמחור"),
    _field("included_seats", "מושבים כלולים", "number", category="תמחור"),
    _field("setup_fee_ils", "דמי הקמה", "number", category="תמחור"),
    _field("override_base_price_ils", "מחיר בסיס מותאם", "number", category="תמחור"),
    _field("override_per_seat_ils", "מחיר למושב מותאם", "number", category="תמחור"),
    _field("override_setup_fee_ils", "דמי הקמה מותאמים", "number", category="תמחור"),
    _field("override_included_seats", "מושבים כלולים מותאמים", "number", category="תמחור"),
    _field("price_lock_reason", "סיבת נעילת מחיר", "string", category="תמחור"),
    _field("notes", "הערות שיוך", "string", category="שיוך מודול"),
    _field("valid_from", "שיוך מתאריך", "date", category="שיוך מודול", groupable=True),
    _field("valid_to", "שיוך עד תאריך", "date", category="שיוך מודול"),
    _field("assignment_created_by", "שיוך נוצר על ידי", "uuid", category="שיוך מודול"),
    _field("assignment_created_at", "שיוך נוצר בתאריך", "datetime", category="שיוך מודול"),
    _field("assignment_updated_by", "שיוך עודכן על ידי", "uuid", category="שיוך מודול"),
    _field("assignment_updated_at", "שיוך עודכן בתאריך", "datetime", category="שיוך מודול"),
    _field("next_renewal_at", "חידוש הבא", "date", category="מנוי", groupable=True),
]

MODULE_CATALOG_FIELDS = [
    _field("module_id", "מזהה מודול", "uuid", category="מודול"),
    _field("module_slug", "קוד מודול", "string", category="מודול", groupable=True),
    _field("module_name", "שם מודול", "string", category="מודול", groupable=True),
    _field("module_description", "תיאור", "string", category="מודול"),
    _field("module_icon", "אייקון", "string", category="מודול"),
    _field("module_color_hex", "צבע", "string", category="מודול", groupable=True),
    _field("module_is_required", "חובה", "boolean", category="מודול", groupable=True),
    _field("module_is_active", "פעיל", "boolean", category="מודול", groupable=True),
    _field("module_sort_order", "סדר תצוגה", "number", category="מודול"),
    _field("module_depends_on", "תלוי ב", "string", category="מודול"),
]

MODULE_PRICING_FIELDS = [
    _field("module_price_id", "מזהה מחיר מודול", "uuid", category="מחיר"),
    _field("module_slug", "קוד מודול", "string", category="מודול", groupable=True),
    _field("module_name", "שם מודול", "string", category="מודול", groupable=True),
    _field("base_price_ils", "מחיר בסיס", "number", category="מחיר"),
    _field("per_seat_ils", "מחיר למושב", "number", category="מחיר"),
    _field("included_seats", "מושבים כלולים", "number", category="מחיר"),
    _field("setup_fee_ils", "דמי הקמה", "number", category="מחיר"),
    _field("valid_from", "מחיר מתאריך", "date", category="מחיר", groupable=True),
    _field("valid_to", "מחיר עד תאריך", "date", category="מחיר"),
    _field("created_by", "מחיר נוצר על ידי", "uuid", category="מחיר"),
    _field("created_at", "מחיר נוצר בתאריך", "datetime", category="מחיר"),
]

TEMPLATE_CATALOG_FIELDS = [
    _field("template_id", "מזהה תבנית", "uuid", category="תבנית"),
    _field("template_name", "שם תבנית", "string", category="תבנית", groupable=True),
    _field("template_description", "תיאור", "string", category="תבנית"),
    _field("default_billing_cycle", "מחזור חיוב ברירת מחדל", "string", category="תבנית", groupable=True),
    _field("trial_days", "ימי ניסיון", "number", category="תבנית"),
    _field("is_active", "פעילה", "boolean", category="תבנית", groupable=True),
    _field("sort_order", "סדר תצוגה", "number", category="תבנית"),
    _field("target_industry", "ענף יעד", "string", category="תבנית", groupable=True),
    _field("recommended_size", "גודל מומלץ", "string", category="תבנית", groupable=True),
    _field("valid_from", "תבנית מתאריך", "date", category="תבנית"),
    _field("valid_to", "תבנית עד תאריך", "date", category="תבנית"),
    _field("created_at", "נוצרה בתאריך", "datetime", category="תבנית"),
    _field("module_count", "כמות מודולים", "number", category="תבנית"),
    _field("default_count", "כמות ברירות מחדל", "number", category="תבנית"),
]

TEMPLATE_MODULE_FIELDS = [
    _field("template_id", "מזהה תבנית", "uuid", category="תבנית"),
    _field("template_name", "שם תבנית", "string", category="תבנית", groupable=True),
    _field("module_slug", "קוד מודול", "string", category="מודול", groupable=True),
    _field("module_name", "שם מודול", "string", category="מודול", groupable=True),
    _field("module_is_required", "מודול חובה", "boolean", category="מודול", groupable=True),
    _field("module_is_active", "מודול פעיל", "boolean", category="מודול", groupable=True),
]

TEMPLATE_DEFAULT_FIELDS = [
    _field("template_default_id", "מזהה ברירת מחדל", "uuid", category="ברירת מחדל"),
    _field("template_id", "מזהה תבנית", "uuid", category="תבנית"),
    _field("template_name", "שם תבנית", "string", category="תבנית", groupable=True),
    _field("default_type", "סוג ברירת מחדל", "string", category="ברירת מחדל", groupable=True),
    _field("default_value", "ערך", "string", category="ברירת מחדל"),
    _field("is_mandatory", "חובה", "boolean", category="ברירת מחדל", groupable=True),
    _field("note", "הערה", "string", category="ברירת מחדל"),
]

ADMIN_USER_FIELDS = [
    _field("admin_user_id", "מזהה משתמש", "uuid", category="משתמש"),
    _field("full_name", "שם מלא", "string", category="משתמש", groupable=True),
    _field("email", "דוא\"ל", "string", category="משתמש"),
    _field("role", "תפקיד", "string", category="משתמש", groupable=True),
    _field("is_active", "פעיל", "boolean", category="משתמש", groupable=True),
    _field("last_login_at", "כניסה אחרונה", "datetime", category="משתמש"),
    _field("created_by", "נוצר על ידי", "uuid", category="משתמש"),
    _field("created_by_name", "נוצר על ידי (שם)", "string", category="משתמש", groupable=True),
    _field("created_at", "נוצר בתאריך", "datetime", category="משתמש"),
    _field("valid_from", "משתמש מתאריך", "date", category="משתמש"),
    _field("valid_to", "משתמש עד תאריך", "date", category="משתמש"),
]

ADMIN_PERMISSION_FIELDS = [
    _field("permission_id", "מזהה הרשאה", "uuid", category="הרשאה"),
    _field("user_id", "מזהה משתמש", "uuid", category="הרשאה"),
    _field("user_name", "שם משתמש", "string", category="הרשאה", groupable=True),
    _field("user_email", "דוא\"ל משתמש", "string", category="הרשאה"),
    _field("resource", "משאב", "string", category="הרשאה", groupable=True),
    _field("can_view", "יכול לצפות", "boolean", category="הרשאה", groupable=True),
    _field("can_edit", "יכול לערוך", "boolean", category="הרשאה", groupable=True),
]

AUDIT_FIELDS = [
    _field("audit_id", "מזהה אירוע", "uuid", category="Audit"),
    _field("audit_tenant_id", "מזהה לקוח", "uuid", category="Audit"),
    _field("audit_tenant_name", "שם לקוח", "string", category="Audit", groupable=True),
    _field("audit_actor_id", "מזהה מבצע", "uuid", category="Audit"),
    _field("audit_actor_name", "שם מבצע", "string", category="Audit", groupable=True),
    _field("audit_actor_email", "דוא\"ל מבצע", "string", category="Audit"),
    _field("audit_actor_type", "סוג מבצע", "string", category="Audit", groupable=True),
    _field("audit_action", "פעולה", "string", category="Audit", groupable=True),
    _field("audit_entity_type", "סוג ישות", "string", category="Audit", groupable=True),
    _field("audit_entity_id", "מזהה ישות", "uuid", category="Audit"),
    _field("audit_old_values", "ערכים קודמים", "string", category="Audit"),
    _field("audit_new_values", "ערכים חדשים", "string", category="Audit"),
    _field("audit_ip_address", "כתובת IP", "string", category="Audit"),
    _field("audit_created_at", "זמן אירוע", "datetime", category="Audit"),
]

SAVED_REPORT_FIELDS = [
    _field("saved_report_id", "מזהה דוח שמור", "uuid", category="דוח שמור"),
    _field("saved_report_name", "שם דוח", "string", category="דוח שמור", groupable=True),
    _field("saved_report_description", "תיאור", "string", category="דוח שמור"),
    _field("saved_report_dataset", "Dataset", "string", category="דוח שמור", groupable=True),
    _field("saved_report_model_dataset", "Dataset בטבלה", "string", category="דוח שמור", groupable=True),
    _field("saved_report_visibility", "נראות", "string", category="דוח שמור", groupable=True),
    _field("saved_report_owner_id", "מזהה בעלים", "uuid", category="דוח שמור"),
    _field("saved_report_owner_name", "בעלים", "string", category="דוח שמור", groupable=True),
    _field("saved_report_definition_json", "הגדרת JSON", "string", category="דוח שמור"),
    _field("saved_report_columns_count", "כמות עמודות", "number", category="דוח שמור"),
    _field("saved_report_filters_count", "כמות סינונים", "number", category="דוח שמור"),
    _field("saved_report_group_by_count", "כמות קיבוצים", "number", category="דוח שמור"),
    _field("saved_report_metrics_count", "כמות מדדים", "number", category="דוח שמור"),
    _field("saved_report_created_at", "נוצר בתאריך", "datetime", category="דוח שמור"),
    _field("saved_report_updated_at", "עודכן בתאריך", "datetime", category="דוח שמור"),
]

MODULE_SUMMARY_FIELDS = [
    _field("module_slug", "קוד מודול", "string", category="מודול", groupable=True),
    _field("module_name", "שם מודול", "string", category="מודול", groupable=True),
    _field("tenant_count", "כמות לקוחות", "number", category="סיכום"),
    _field("total_seats", "סה\"כ מושבים", "number", category="סיכום"),
    _field("avg_seats", "ממוצע מושבים", "number", category="סיכום"),
    _field("first_assigned_at", "שיוך ראשון", "date", category="סיכום", groupable=True),
    _field("last_assigned_at", "שיוך אחרון", "date", category="סיכום", groupable=True),
]

SEAT_DISTRIBUTION_FIELDS = [
    _field("seat_bucket", "טווח מושבים", "string", category="סיכום", groupable=True),
    _field("tenant_count", "כמות לקוחות", "number", category="סיכום"),
    _field("total_seats", "סה\"כ מושבים", "number", category="סיכום"),
    _field("avg_seats", "ממוצע מושבים", "number", category="סיכום"),
]

def _merge_fields(*groups: list[ReportFieldDefinition]) -> list[ReportFieldDefinition]:
    merged: dict[str, ReportFieldDefinition] = {}
    for group in groups:
        for field in group:
            if field.id not in merged:
                merged[field.id] = field
    return list(merged.values())


# Period / range mode fields — injected into datasets when date_from+date_to are set
PERIOD_FIELDS = [
    _field("period_valid_from", "תחילת תקופה", "date", category="תקופה", groupable=True,
           description="תאריך תחילת התוקף שבו הרשומה היתה פעילה"),
    _field("period_valid_to", "סיום תקופה", "date", category="תקופה", groupable=True,
           description="תאריך סוף התוקף שבו הרשומה היתה פעילה (ריק = פעיל עד עיד)"),
    _field("period_duration_days", "אורך תקופה (ימים)", "number", category="תקופה",
           description="כמה ימים הרשומה היתה פעילה בתוך הטווח הנבחר"),
]
PERIOD_FIELD_IDS = {f.id for f in PERIOD_FIELDS}

MASTER_FIELDS = _merge_fields(
    PERIOD_FIELDS,
    [
        _field("record_type", "סוג רשומה", "string", category="כללי", groupable=True, description="מאיזה מקור נתונים הגיעה הרשומה"),
        _field("record_key", "מפתח רשומה", "string", category="כללי", description="מזהה פנימי של הרשומה המאוחדת"),
        _field("master_created_at", "תאריך רשומה ראשי", "datetime", category="כללי", groupable=True),
        _field("tenant_org_number", "מספר ארגון", "number", category="לקוח", groupable=True),
        _field("admin_full_name", "שם משתמש אדמין", "string", category="משתמש", groupable=True),
        _field("permission_resource", "משאב הרשאה", "string", category="הרשאה", groupable=True),
        _field("audit_action", "פעולת Audit", "string", category="Audit", groupable=True),
        _field("saved_report_name", "שם דוח שמור", "string", category="דוח שמור", groupable=True),
    ],
    TENANT_SNAPSHOT_FIELDS,
    TENANT_MODULE_FIELDS,
    MODULE_CATALOG_FIELDS,
    MODULE_PRICING_FIELDS,
    TEMPLATE_CATALOG_FIELDS,
    TEMPLATE_MODULE_FIELDS,
    TEMPLATE_DEFAULT_FIELDS,
    ADMIN_USER_FIELDS,
    ADMIN_PERMISSION_FIELDS,
    AUDIT_FIELDS,
    SAVED_REPORT_FIELDS,
)


def _dataset(
    dataset_id: str,
    label: str,
    description: str,
    fields: list[ReportFieldDefinition],
    default_columns: list[str],
    metrics: list[ReportMetricDefinition],
) -> ReportDatasetDefinition:
    enriched_fields = _enrich_dataset_fields(fields)
    return ReportDatasetDefinition(
        id=dataset_id,
        label=label,
        description=description,
        fields=enriched_fields,
        default_columns=default_columns,
        groupable_fields=[field.id for field in enriched_fields if field.groupable],
        metrics=metrics,
    )


DATASETS = [
    _dataset(
        "master_dataset",
        "מרכז נתונים מאוחד",
        "מאגר־על שמרכז את כל שדות הליבה במערכת, כולל כתובות, טלפונים, אנשי קשר, מנויים, מודולים, תבניות, אדמינים ו-Audit.",
        MASTER_FIELDS,
        [
            "record_type",
            "org_number",
            "identity_name_he",
            "contact_main_name",
            "contact_main_phone",
            "contact_main_email",
            "address_main_city",
            "address_main_street",
            "module_name",
            "template_name",
        ],
        [ReportMetricDefinition(operation="count", label="כמות רשומות")],
    ),
    _dataset(
        "tenant_snapshot_full",
        "לקוחות",
        "שורה אחת לכל לקוח עם זהות, קשר, כתובת, סטטוס ומנוי פעיל במועד הדוח.",
        TENANT_SNAPSHOT_FIELDS,
        ["org_number", "identity_name_he", "status_value", "subscription_seat_count", "module_count", "subscription_next_renewal_at"],
        [
            ReportMetricDefinition(operation="count", label="לקוחות"),
            ReportMetricDefinition(operation="sum", field="subscription_seat_count", label='סה"כ מושבים'),
            ReportMetricDefinition(operation="avg", field="subscription_discount_pct", label="ממוצע הנחה"),
        ],
    ),
    _dataset(
        "tenant_module_snapshot_full",
        "לקוחות לפי מודול",
        "שורה לכל שיוך מודול ללקוח עם פרטי מודול, מנוי ותמחור.",
        TENANT_MODULE_FIELDS,
        ["tenant_name", "module_name", "module_seats", "tenant_status", "valid_from", "next_renewal_at"],
        [
            ReportMetricDefinition(operation="count", label="שיוכי מודול"),
            ReportMetricDefinition(operation="count_distinct", field="tenant_id", label="לקוחות"),
            ReportMetricDefinition(operation="sum", field="module_seats", label='סה"כ מושבי מודול'),
            ReportMetricDefinition(operation="sum", field="base_price_ils", label='סה"כ מחיר בסיס'),
        ],
    ),
    _dataset(
        "module_catalog",
        "קטלוג מודולים",
        "מודולים פעילים ולא פעילים עם כל שדות הקטלוג וההגדרה.",
        MODULE_CATALOG_FIELDS,
        ["module_slug", "module_name", "module_is_active", "module_is_required", "module_sort_order"],
        [ReportMetricDefinition(operation="count", label="מודולים")],
    ),
    _dataset(
        "module_pricing",
        "מחירי מודולים",
        "תמחור מודולים אפקטיבי במועד הדוח.",
        MODULE_PRICING_FIELDS,
        ["module_name", "base_price_ils", "per_seat_ils", "included_seats", "setup_fee_ils", "valid_from"],
        [
            ReportMetricDefinition(operation="count", label="רשומות מחיר"),
            ReportMetricDefinition(operation="sum", field="base_price_ils", label='סה"כ מחיר בסיס'),
        ],
    ),
    _dataset(
        "template_catalog",
        "תבניות",
        "קטלוג התבניות, הגדרות ברירת המחדל וכיסוי מודולים.",
        TEMPLATE_CATALOG_FIELDS,
        ["template_name", "default_billing_cycle", "trial_days", "module_count", "default_count", "is_active"],
        [ReportMetricDefinition(operation="count", label="תבניות")],
    ),
    _dataset(
        "template_modules",
        "שיוכי מודולים לתבניות",
        "הקשר בין תבניות למודולים.",
        TEMPLATE_MODULE_FIELDS,
        ["template_name", "module_name", "module_is_required", "module_is_active"],
        [ReportMetricDefinition(operation="count", label="שיוכי תבנית-מודול")],
    ),
    _dataset(
        "template_defaults",
        "ברירות מחדל של תבניות",
        "ערכי ברירת מחדל לכל תבנית.",
        TEMPLATE_DEFAULT_FIELDS,
        ["template_name", "default_type", "default_value", "is_mandatory"],
        [ReportMetricDefinition(operation="count", label="ברירות מחדל")],
    ),
    _dataset(
        "admin_users",
        "משתמשי אדמין",
        "משתמשי הניהול במערכת וכל שדות הזהות והסטטוס שלהם.",
        ADMIN_USER_FIELDS,
        ["full_name", "email", "role", "is_active", "last_login_at"],
        [ReportMetricDefinition(operation="count", label="משתמשי אדמין")],
    ),
    _dataset(
        "admin_permissions",
        "הרשאות אדמין",
        "שורה לכל הרשאת משאב של משתמש אדמין.",
        ADMIN_PERMISSION_FIELDS,
        ["user_name", "resource", "can_view", "can_edit"],
        [ReportMetricDefinition(operation="count", label="הרשאות")],
    ),
    _dataset(
        "audit_logs",
        "Audit Log",
        "לוג פעולות מערכת עם שחקן, ישות ושינוי.",
        AUDIT_FIELDS,
        ["audit_created_at", "audit_actor_name", "audit_action", "audit_entity_type", "audit_tenant_name"],
        [ReportMetricDefinition(operation="count", label="אירועי Audit")],
    ),
    _dataset(
        "saved_reports",
        "דוחות שמורים",
        "הגדרות הדוחות השמורים של המשתמשים במערכת.",
        SAVED_REPORT_FIELDS,
        ["saved_report_name", "saved_report_dataset", "saved_report_visibility", "saved_report_owner_name", "saved_report_updated_at"],
        [ReportMetricDefinition(operation="count", label="דוחות שמורים")],
    ),
    _dataset(
        "module_summary",
        "סיכום מודולים",
        "תמונת סיכום של אימוץ מודולים לפי כמות לקוחות ומושבים.",
        MODULE_SUMMARY_FIELDS,
        ["module_name", "tenant_count", "total_seats", "avg_seats", "last_assigned_at"],
        [
            ReportMetricDefinition(operation="count", label="מודולים"),
            ReportMetricDefinition(operation="sum", field="tenant_count", label='סה"כ לקוחות'),
            ReportMetricDefinition(operation="sum", field="total_seats", label='סה"כ מושבים'),
        ],
    ),
    _dataset(
        "seat_distribution",
        "התפלגות מושבים",
        "פילוח לקוחות לפי טווחי מושבים במנוי.",
        SEAT_DISTRIBUTION_FIELDS,
        ["seat_bucket", "tenant_count", "total_seats", "avg_seats"],
        [
            ReportMetricDefinition(operation="count", label="קבוצות"),
            ReportMetricDefinition(operation="sum", field="tenant_count", label='סה"כ לקוחות'),
            ReportMetricDefinition(operation="sum", field="total_seats", label='סה"כ מושבים'),
        ],
    ),
]

DATASET_MAP = {dataset.id: dataset for dataset in DATASETS}
CATALOG = [
    ReportCatalogItem(
        id="tenant_portfolio",
        title="פורטפוליו לקוחות",
        description="תמונת מצב מלאה של לקוחות, מנוי, סטטוס וחידוש.",
        dataset="tenant_snapshot_full",
        definition=ReportDefinition(
            dataset="tenant_snapshot_full",
            columns=["org_number", "identity_name_he", "status_value", "subscription_seat_count", "module_count", "subscription_next_renewal_at"],
            sort=[{"field": "subscription_next_renewal_at", "direction": "asc"}],
            metrics=[
                {"operation": "count", "label": "לקוחות"},
                {"operation": "sum", "field": "subscription_seat_count", "label": 'סה"כ מושבים'},
            ],
        ),
    ),
    ReportCatalogItem(
        id="customers_by_module",
        title="לקוחות לפי מודול",
        description="אילו לקוחות מחזיקים כל מודול, מאיזה תאריך ובכמה מושבים.",
        dataset="tenant_module_snapshot_full",
        definition=ReportDefinition(
            dataset="tenant_module_snapshot_full",
            columns=["tenant_name", "module_name", "module_seats", "pricing_mode", "valid_from", "tenant_status"],
            sort=[{"field": "valid_from", "direction": "desc"}],
            metrics=[
                {"operation": "count_distinct", "field": "tenant_id", "label": "לקוחות"},
                {"operation": "sum", "field": "module_seats", "label": 'סה"כ מושבי מודול'},
            ],
        ),
    ),
    ReportCatalogItem(
        id="module_pricing_watch",
        title="תמחור מודולים",
        description="השוואת מחירי קטלוג ודמי הקמה לפי מודול.",
        dataset="module_pricing",
        definition=ReportDefinition(
            dataset="module_pricing",
            columns=["module_name", "base_price_ils", "per_seat_ils", "included_seats", "setup_fee_ils", "valid_from"],
            sort=[{"field": "module_name", "direction": "asc"}],
            metrics=[{"operation": "count", "label": "רשומות מחיר"}],
        ),
    ),
    ReportCatalogItem(
        id="admin_access_matrix",
        title="מטריצת הרשאות אדמין",
        description="מי יכול לצפות ולערוך בכל משאב ניהולי.",
        dataset="admin_permissions",
        definition=ReportDefinition(
            dataset="admin_permissions",
            columns=["user_name", "user_email", "resource", "can_view", "can_edit"],
            sort=[{"field": "user_name", "direction": "asc"}],
            metrics=[{"operation": "count", "label": "הרשאות"}],
        ),
    ),
    ReportCatalogItem(
        id="audit_watchlist",
        title="Audit Watchlist",
        description="מעקב אחר פעולות מערכת לפי מבצע, לקוח וישות.",
        dataset="audit_logs",
        definition=ReportDefinition(
            dataset="audit_logs",
            columns=["audit_created_at", "audit_actor_name", "audit_action", "audit_entity_type", "audit_tenant_name"],
            sort=[{"field": "audit_created_at", "direction": "desc"}],
            metrics=[{"operation": "count", "label": "אירועים"}],
        ),
    ),
]


def _effective(record: Any, as_of: date) -> bool:
    valid_from = getattr(record, "valid_from", None)
    valid_to = getattr(record, "valid_to", None)
    if valid_from and valid_from > as_of:
        return False
    if valid_to and valid_to <= as_of:
        return False
    return True


def _pick_temporal(records: list[Any], key_attr: str, as_of: date) -> dict[Any, Any]:
    grouped: dict[Any, list[Any]] = defaultdict(list)
    for record in records:
        if _effective(record, as_of):
            grouped[getattr(record, key_attr)].append(record)
    selected: dict[Any, Any] = {}
    for key, items in grouped.items():
        selected[key] = max(items, key=lambda item: getattr(item, "valid_from", date.min) or date.min)
    return selected


def _join_text(values: list[Any] | None) -> str | None:
    if not values:
        return None
    return ", ".join(str(value) for value in values if value not in (None, ""))


def _resolve_user_label(users: dict[uuid.UUID, AdminUser], user_id: uuid.UUID | None) -> str | None:
    if user_id is None:
        return None
    user = users.get(user_id)
    if user is None:
        return str(user_id)
    return user.full_name or user.email


def _normalize_dataset_id(dataset_id: str) -> str:
    return LEGACY_DATASET_ALIASES.get(dataset_id, dataset_id)


def _normalize_definition(definition: ReportDefinition) -> ReportDefinition:
    normalized_dataset = _normalize_dataset_id(definition.dataset)
    if normalized_dataset == definition.dataset:
        return definition
    return definition.model_copy(update={"dataset": normalized_dataset})


async def _load_filter_options(db: AsyncSession) -> ReportFilterOptions:
    module_result = await db.execute(select(Module).order_by(Module.sort_order, Module.name))
    modules = module_result.scalars().all()
    return ReportFilterOptions(
        tenant_statuses=TENANT_STATUS_OPTIONS,
        modules=[ReportFilterOption(value=module.slug, label=module.name) for module in modules],
    )


async def get_catalog(db: AsyncSession) -> ReportCatalogResponse:
    return ReportCatalogResponse(reports=CATALOG, filter_options=await _load_filter_options(db))


async def get_datasets(db: AsyncSession) -> ReportDatasetsResponse:
    return ReportDatasetsResponse(datasets=DATASETS, filter_options=await _load_filter_options(db))


async def _load_snapshot_rows(db: AsyncSession, as_of: date) -> dict[str, list[dict[str, Any]]]:
    tenant_result = await db.execute(select(Tenant))
    identity_result = await db.execute(select(TenantIdentity))
    contact_result = await db.execute(select(TenantContact).where(TenantContact.contact_type == "main"))
    address_result = await db.execute(select(TenantAddress).where(TenantAddress.addr_type == "main"))
    status_result = await db.execute(select(TenantStatus))
    subscription_result = await db.execute(select(TenantSubscription))
    subscription_module_result = await db.execute(select(TenantSubscriptionModule))
    module_result = await db.execute(select(Module))
    module_price_result = await db.execute(select(ModulePrice))
    template_result = await db.execute(select(OrgTemplate))
    template_default_result = await db.execute(select(OrgTemplateDefault))
    template_module_result = await db.execute(select(OrgTemplateModule))
    admin_user_result = await db.execute(select(AdminUser))
    admin_permission_result = await db.execute(select(AdminUserPermission))
    audit_result = await db.execute(select(AuditLog))
    try:
        saved_report_result = await db.execute(select(SavedReportView))
        saved_reports = saved_report_result.scalars().all()
    except ProgrammingError as exc:
        if "saved_report_views" in str(exc).lower():
            saved_reports = []
        else:
            raise

    tenants = tenant_result.scalars().all()
    identities = _pick_temporal(identity_result.scalars().all(), "tenant_id", as_of)
    contacts = _pick_temporal(contact_result.scalars().all(), "tenant_id", as_of)
    addresses = _pick_temporal(address_result.scalars().all(), "tenant_id", as_of)
    statuses = _pick_temporal(status_result.scalars().all(), "tenant_id", as_of)
    subscriptions = _pick_temporal(subscription_result.scalars().all(), "tenant_id", as_of)
    modules = {row.slug: row for row in module_result.scalars().all()}
    prices = _pick_temporal(module_price_result.scalars().all(), "module_slug", as_of)
    templates = {row.id: row for row in template_result.scalars().all()}
    template_defaults = template_default_result.scalars().all()
    template_modules = template_module_result.scalars().all()
    admin_users = admin_user_result.scalars().all()
    admin_user_lookup = {row.id: row for row in admin_users}
    permissions = admin_permission_result.scalars().all()
    audit_logs = audit_result.scalars().all()

    module_rows_by_subscription: dict[uuid.UUID, list[TenantSubscriptionModule]] = defaultdict(list)
    effective_modules: dict[tuple[uuid.UUID, str], TenantSubscriptionModule] = {}
    for row in subscription_module_result.scalars().all():
        if not _effective(row, as_of):
            continue
        key = (row.tenant_subscription_id, row.module_slug)
        current = effective_modules.get(key)
        if current is None or (current.valid_from or date.min) < (row.valid_from or date.min):
            effective_modules[key] = row
    for row in effective_modules.values():
        module_rows_by_subscription[row.tenant_subscription_id].append(row)

    template_default_counts: dict[uuid.UUID, int] = defaultdict(int)
    for row in template_defaults:
        template_default_counts[row.template_id] += 1

    template_module_counts: dict[uuid.UUID, int] = defaultdict(int)
    for row in template_modules:
        template_module_counts[row.template_id] += 1

    tenant_name_by_id: dict[uuid.UUID, str] = {}
    tenant_snapshot_rows: list[dict[str, Any]] = []
    tenant_module_rows: list[dict[str, Any]] = []
    master_rows: list[dict[str, Any]] = []

    for tenant in tenants:
        identity = identities.get(tenant.tenant_id)
        contact = contacts.get(tenant.tenant_id)
        address = addresses.get(tenant.tenant_id)
        status = statuses.get(tenant.tenant_id)
        subscription = subscriptions.get(tenant.tenant_id)
        subscription_modules = module_rows_by_subscription.get(subscription.id, []) if subscription else []
        derived_seat_count, derived_module_slugs = derive_subscription_snapshot(subscription_modules)
        module_names = [
            modules[row.module_slug].name if row.module_slug in modules else row.module_slug
            for row in subscription_modules
        ]
        template = templates.get(getattr(subscription, "template_id", None)) if subscription else None
        tenant_name = getattr(identity, "name_he", None) or f"Tenant {tenant.org_number}"
        tenant_name_by_id[tenant.tenant_id] = tenant_name

        snapshot_row = {
            "tenant_id": tenant.tenant_id,
            "org_number": tenant.org_number,
            "tenant_created_at": tenant.created_at,
            "identity_id": getattr(identity, "id", None),
            "identity_name_he": tenant_name,
            "identity_name_en": getattr(identity, "name_en", None),
            "identity_tax_id": getattr(identity, "tax_id", None),
            "identity_entity_type": getattr(identity, "entity_type", None),
            "identity_logo_url": getattr(identity, "logo_url", None),
            "identity_industry_code": getattr(identity, "industry_code", None),
            "identity_valid_from": getattr(identity, "valid_from", None),
            "identity_valid_to": getattr(identity, "valid_to", None),
            "identity_created_by": getattr(identity, "created_by", None),
            "identity_created_at": getattr(identity, "created_at", None),
            "identity_updated_by": getattr(identity, "updated_by", None),
            "identity_updated_at": getattr(identity, "updated_at", None),
            "contact_main_id": getattr(contact, "id", None),
            "contact_main_type": getattr(contact, "contact_type", None),
            "contact_main_name": getattr(contact, "contact_name", None),
            "contact_main_email": getattr(contact, "email", None),
            "contact_main_phone": getattr(contact, "phone", None),
            "contact_main_phone_alt": getattr(contact, "phone_alt", None),
            "contact_main_website": getattr(contact, "website", None),
            "contact_main_valid_from": getattr(contact, "valid_from", None),
            "contact_main_valid_to": getattr(contact, "valid_to", None),
            "contact_main_created_by": getattr(contact, "created_by", None),
            "contact_main_created_at": getattr(contact, "created_at", None),
            "contact_main_updated_by": getattr(contact, "updated_by", None),
            "contact_main_updated_at": getattr(contact, "updated_at", None),
            "address_main_id": getattr(address, "id", None),
            "address_main_type": getattr(address, "addr_type", None),
            "address_main_street": getattr(address, "street", None),
            "address_main_city": getattr(address, "city", None),
            "address_main_zip_code": getattr(address, "zip_code", None),
            "address_main_country": getattr(address, "country", None),
            "address_main_valid_from": getattr(address, "valid_from", None),
            "address_main_valid_to": getattr(address, "valid_to", None),
            "address_main_created_by": getattr(address, "created_by", None),
            "address_main_created_at": getattr(address, "created_at", None),
            "address_main_updated_by": getattr(address, "updated_by", None),
            "address_main_updated_at": getattr(address, "updated_at", None),
            "status_id": getattr(status, "id", None),
            "status_value": getattr(status, "status", None),
            "status_reason": getattr(status, "reason", None),
            "status_notes": getattr(status, "notes", None),
            "status_valid_from": getattr(status, "valid_from", None),
            "status_valid_to": getattr(status, "valid_to", None),
            "status_created_by": getattr(status, "created_by", None),
            "status_created_at": getattr(status, "created_at", None),
            "status_updated_by": getattr(status, "updated_by", None),
            "status_updated_at": getattr(status, "updated_at", None),
            "subscription_id": getattr(subscription, "id", None),
            "subscription_billing_cycle": getattr(subscription, "billing_cycle", None),
            "subscription_currency": getattr(subscription, "currency", None),
            "subscription_template_id": getattr(subscription, "template_id", None),
            "subscription_template_name": getattr(template, "name", None),
            "subscription_seat_count": derived_seat_count,
            "subscription_selected_module_slugs": _join_text(derived_module_slugs),
            "subscription_discount_pct": float(getattr(subscription, "discount_pct", 0) or 0),
            "subscription_is_price_locked": getattr(subscription, "is_price_locked", False),
            "subscription_next_renewal_at": getattr(subscription, "next_renewal_at", None),
            "subscription_valid_from": getattr(subscription, "valid_from", None),
            "subscription_valid_to": getattr(subscription, "valid_to", None),
            "subscription_created_by": getattr(subscription, "created_by", None),
            "subscription_created_at": getattr(subscription, "created_at", None),
            "subscription_updated_by": getattr(subscription, "updated_by", None),
            "subscription_updated_at": getattr(subscription, "updated_at", None),
            "module_count": len(subscription_modules),
            "module_names": _join_text(module_names),
        }
        tenant_snapshot_rows.append(snapshot_row)
        master_rows.append(
            {
                "record_type": "tenant_snapshot",
                "record_key": str(tenant.tenant_id),
                "master_created_at": tenant.created_at,
                "tenant_org_number": tenant.org_number,
                **snapshot_row,
            }
        )

        for row in subscription_modules:
            module = modules.get(row.module_slug)
            price = prices.get(row.module_slug)
            module_row = {
                "tenant_id": tenant.tenant_id,
                "org_number": tenant.org_number,
                "tenant_name": tenant_name,
                "tenant_status": getattr(status, "status", None),
                "subscription_id": getattr(subscription, "id", None),
                "subscription_template_name": getattr(template, "name", None),
                "subscription_billing_cycle": getattr(subscription, "billing_cycle", None),
                "subscription_currency": getattr(subscription, "currency", None),
                "subscription_seat_count": derived_seat_count,
                "subscription_created_by": getattr(subscription, "created_by", None),
                "subscription_created_at": getattr(subscription, "created_at", None),
                "subscription_updated_by": getattr(subscription, "updated_by", None),
                "subscription_updated_at": getattr(subscription, "updated_at", None),
                "module_assignment_id": row.id,
                "tenant_subscription_id": row.tenant_subscription_id,
                "module_slug": row.module_slug,
                "module_name": getattr(module, "name", row.module_slug),
                "module_description": getattr(module, "description", None),
                "module_icon": getattr(module, "icon", None),
                "module_color_hex": getattr(module, "color_hex", None),
                "module_is_required": getattr(module, "is_required", None),
                "module_is_active": getattr(module, "is_active", None),
                "module_sort_order": getattr(module, "sort_order", None),
                "module_depends_on": _join_text(getattr(module, "depends_on", None)),
                "source_type": row.source_type,
                "module_status": row.status,
                "module_seats": int(row.seats or 0),
                "pricing_mode": row.pricing_mode,
                "base_price_ils": float(getattr(price, "base_price_ils", 0) or 0),
                "per_seat_ils": float(getattr(price, "per_seat_ils", 0) or 0),
                "included_seats": int(getattr(price, "included_seats", 0) or 0),
                "setup_fee_ils": float(getattr(price, "setup_fee_ils", 0) or 0),
                "override_base_price_ils": float(getattr(row, "override_base_price_ils", 0) or 0),
                "override_per_seat_ils": float(getattr(row, "override_per_seat_ils", 0) or 0),
                "override_setup_fee_ils": float(getattr(row, "override_setup_fee_ils", 0) or 0),
                "override_included_seats": int(getattr(row, "override_included_seats", 0) or 0) if getattr(row, "override_included_seats", None) is not None else None,
                "price_lock_reason": getattr(row, "price_lock_reason", None),
                "notes": getattr(row, "notes", None),
                "valid_from": row.valid_from,
                "valid_to": row.valid_to,
                "assignment_created_by": getattr(row, "created_by", None),
                "assignment_created_at": getattr(row, "created_at", None),
                "assignment_updated_by": getattr(row, "updated_by", None),
                "assignment_updated_at": getattr(row, "updated_at", None),
                "next_renewal_at": getattr(subscription, "next_renewal_at", None),
            }
            tenant_module_rows.append(module_row)
            master_rows.append(
                {
                    "record_type": "tenant_module",
                    "record_key": str(row.id),
                    "master_created_at": row.valid_from or tenant.created_at,
                    "tenant_org_number": tenant.org_number,
                    **snapshot_row,
                    **module_row,
                }
            )

    module_catalog_rows: list[dict[str, Any]] = []
    module_pricing_rows: list[dict[str, Any]] = []
    for module in modules.values():
        module_catalog_rows.append(
            {
                "module_id": module.id,
                "module_slug": module.slug,
                "module_name": module.name,
                "module_description": module.description,
                "module_icon": module.icon,
                "module_color_hex": module.color_hex,
                "module_is_required": module.is_required,
                "module_is_active": module.is_active,
                "module_sort_order": module.sort_order,
                "module_depends_on": _join_text(module.depends_on),
            }
        )
        master_rows.append(
            {
                "record_type": "module",
                "record_key": module.slug,
                **module_catalog_rows[-1],
            }
        )
    for slug, price in prices.items():
        module = modules.get(slug)
        module_pricing_rows.append(
            {
                "module_price_id": getattr(price, "id", None),
                "module_slug": slug,
                "module_name": getattr(module, "name", slug),
                "base_price_ils": float(getattr(price, "base_price_ils", 0) or 0),
                "per_seat_ils": float(getattr(price, "per_seat_ils", 0) or 0),
                "included_seats": int(getattr(price, "included_seats", 0) or 0),
                "setup_fee_ils": float(getattr(price, "setup_fee_ils", 0) or 0),
                "valid_from": getattr(price, "valid_from", None),
                "valid_to": getattr(price, "valid_to", None),
                "created_by": getattr(price, "created_by", None),
                "created_at": getattr(price, "created_at", None),
            }
        )
        master_rows.append(
            {
                "record_type": "module_price",
                "record_key": slug,
                "master_created_at": getattr(price, "created_at", None),
                **module_pricing_rows[-1],
            }
        )

    template_catalog_rows: list[dict[str, Any]] = []
    for template in templates.values():
        template_catalog_rows.append(
            {
                "template_id": template.id,
                "template_name": template.name,
                "template_description": template.description,
                "default_billing_cycle": template.default_billing_cycle,
                "trial_days": template.trial_days,
                "is_active": template.is_active,
                "sort_order": template.sort_order,
                "target_industry": template.target_industry,
                "recommended_size": template.recommended_size,
                "valid_from": template.valid_from,
                "valid_to": template.valid_to,
                "created_at": template.created_at,
                "module_count": template_module_counts.get(template.id, 0),
                "default_count": template_default_counts.get(template.id, 0),
            }
        )
        master_rows.append(
            {
                "record_type": "template",
                "record_key": str(template.id),
                "master_created_at": template.created_at,
                **template_catalog_rows[-1],
            }
        )

    template_module_rows: list[dict[str, Any]] = []
    for row in template_modules:
        template = templates.get(row.template_id)
        module = modules.get(row.module_slug)
        template_module_rows.append(
            {
                "template_id": row.template_id,
                "template_name": getattr(template, "name", str(row.template_id)),
                "module_slug": row.module_slug,
                "module_name": getattr(module, "name", row.module_slug),
                "module_is_required": getattr(module, "is_required", None),
                "module_is_active": getattr(module, "is_active", None),
            }
        )
        master_rows.append(
            {
                "record_type": "template_module",
                "record_key": f"{row.template_id}:{row.module_slug}",
                **template_module_rows[-1],
            }
        )

    template_default_rows: list[dict[str, Any]] = []
    for row in template_defaults:
        template = templates.get(row.template_id)
        template_default_rows.append(
            {
                "template_default_id": row.id,
                "template_id": row.template_id,
                "template_name": getattr(template, "name", str(row.template_id)),
                "default_type": row.default_type,
                "default_value": row.default_value,
                "is_mandatory": row.is_mandatory,
                "note": row.note,
            }
        )
        master_rows.append(
            {
                "record_type": "template_default",
                "record_key": str(row.id),
                **template_default_rows[-1],
            }
        )

    admin_user_rows: list[dict[str, Any]] = []
    for user in admin_users:
        admin_user_rows.append(
            {
                "admin_user_id": user.id,
                "full_name": user.full_name,
                "email": user.email,
                "role": user.role,
                "is_active": user.is_active,
                "last_login_at": user.last_login_at,
                "created_by": user.created_by,
                "created_by_name": _resolve_user_label(admin_user_lookup, user.created_by),
                "created_at": user.created_at,
                "valid_from": user.valid_from,
                "valid_to": user.valid_to,
            }
        )
        master_rows.append(
            {
                "record_type": "admin_user",
                "record_key": str(user.id),
                "master_created_at": user.created_at,
                "admin_full_name": user.full_name,
                **admin_user_rows[-1],
            }
        )

    admin_permission_rows: list[dict[str, Any]] = []
    for row in permissions:
        user = admin_user_lookup.get(row.user_id)
        admin_permission_rows.append(
            {
                "permission_id": row.id,
                "user_id": row.user_id,
                "user_name": getattr(user, "full_name", None),
                "user_email": getattr(user, "email", None),
                "resource": row.resource,
                "can_view": row.can_view,
                "can_edit": row.can_edit,
            }
        )
        master_rows.append(
            {
                "record_type": "admin_permission",
                "record_key": str(row.id),
                "admin_full_name": getattr(user, "full_name", None),
                "permission_resource": row.resource,
                **admin_permission_rows[-1],
            }
        )

    audit_rows: list[dict[str, Any]] = []
    for row in audit_logs:
        actor = admin_user_lookup.get(row.actor_id)
        audit_rows.append(
            {
                "audit_id": row.id,
                "audit_tenant_id": row.tenant_id,
                "audit_tenant_name": tenant_name_by_id.get(row.tenant_id) if row.tenant_id else None,
                "audit_actor_id": row.actor_id,
                "audit_actor_name": getattr(actor, "full_name", None),
                "audit_actor_email": getattr(actor, "email", None),
                "audit_actor_type": row.actor_type,
                "audit_action": row.action,
                "audit_entity_type": row.entity_type,
                "audit_entity_id": row.entity_id,
                "audit_old_values": _serialize_value(row.old_values),
                "audit_new_values": _serialize_value(row.new_values),
                "audit_ip_address": row.ip_address,
                "audit_created_at": row.created_at,
            }
        )
        master_rows.append(
            {
                "record_type": "audit_log",
                "record_key": str(row.id),
                "master_created_at": row.created_at,
                "identity_name_he": tenant_name_by_id.get(row.tenant_id) if row.tenant_id else None,
                "audit_action": row.action,
                "audit_actor_email": getattr(actor, "email", None),
                **audit_rows[-1],
            }
        )

    saved_report_rows: list[dict[str, Any]] = []
    for row in saved_reports:
        definition = _normalize_definition(ReportDefinition.model_validate(row.definition_json))
        owner = admin_user_lookup.get(row.owner_id)
        saved_report_rows.append(
            {
                "saved_report_id": row.id,
                "saved_report_name": row.name,
                "saved_report_description": row.description,
                "saved_report_dataset": definition.dataset,
                "saved_report_model_dataset": row.dataset,
                "saved_report_visibility": row.visibility,
                "saved_report_owner_id": row.owner_id,
                "saved_report_owner_name": getattr(owner, "full_name", None),
                "saved_report_definition_json": _serialize_value(row.definition_json),
                "saved_report_columns_count": len(definition.columns),
                "saved_report_filters_count": len(definition.filters),
                "saved_report_group_by_count": len(definition.group_by),
                "saved_report_metrics_count": len(definition.metrics),
                "saved_report_created_at": row.created_at,
                "saved_report_updated_at": row.updated_at,
            }
        )
        master_rows.append(
            {
                "record_type": "saved_report",
                "record_key": str(row.id),
                "master_created_at": row.updated_at or row.created_at,
                "saved_report_name": row.name,
                "saved_report_visibility": row.visibility,
                **saved_report_rows[-1],
            }
        )

    grouped_modules: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in tenant_module_rows:
        grouped_modules[row["module_slug"]].append(row)

    module_summary_rows = []
    for slug, rows in grouped_modules.items():
        seats = [int(item["module_seats"] or 0) for item in rows]
        assigned_dates = [item["valid_from"] for item in rows if item["valid_from"]]
        module_summary_rows.append(
            {
                "module_slug": slug,
                "module_name": rows[0]["module_name"],
                "tenant_count": len({item["tenant_id"] for item in rows}),
                "total_seats": sum(seats),
                "avg_seats": round(sum(seats) / len(seats), 2) if seats else 0,
                "first_assigned_at": min(assigned_dates) if assigned_dates else None,
                "last_assigned_at": max(assigned_dates) if assigned_dates else None,
            }
        )

    seat_distribution_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in tenant_snapshot_rows:
        seat_distribution_groups[_seat_bucket(int(row["subscription_seat_count"] or 0))].append(row)

    seat_distribution_rows = []
    for bucket, rows in seat_distribution_groups.items():
        seats = [int(item["subscription_seat_count"] or 0) for item in rows]
        seat_distribution_rows.append(
            {
                "seat_bucket": bucket,
                "tenant_count": len(rows),
                "total_seats": sum(seats),
                "avg_seats": round(sum(seats) / len(seats), 2) if seats else 0,
            }
        )

    return {
        "master_dataset": master_rows,
        "tenant_snapshot_full": tenant_snapshot_rows,
        "tenant_module_snapshot_full": tenant_module_rows,
        "module_catalog": module_catalog_rows,
        "module_pricing": module_pricing_rows,
        "template_catalog": template_catalog_rows,
        "template_modules": template_module_rows,
        "template_defaults": template_default_rows,
        "admin_users": admin_user_rows,
        "admin_permissions": admin_permission_rows,
        "audit_logs": audit_rows,
        "saved_reports": saved_report_rows,
        "module_summary": module_summary_rows,
        "seat_distribution": seat_distribution_rows,
        "tenant_snapshot": tenant_snapshot_rows,
        "tenant_module_snapshot": tenant_module_rows,
    }


def _normalize_compare_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _coerce_filter_value(field_type: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, list):
        return [_coerce_filter_value(field_type, item) for item in value]
    if field_type == "number":
        if value == "":
            return None
        if "." in str(value):
            return Decimal(str(value))
        return int(value)
    if field_type in {"date", "datetime"} and isinstance(value, str) and value:
        return date.fromisoformat(value[:10])
    if field_type == "uuid" and isinstance(value, str) and value:
        return uuid.UUID(value)
    if field_type == "boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "כן"}
    return value


def _match_filter(row_value: Any, operator: str, expected: Any) -> bool:
    row_value = _normalize_compare_value(row_value)
    expected = _normalize_compare_value(expected)
    if operator == "is_null":
        return row_value is None
    if operator == "is_not_null":
        return row_value is not None
    if operator in {"in", "not_in"} and isinstance(expected, str):
        expected = [part.strip() for part in expected.split(",") if part.strip()]
    if operator == "equals":
        return row_value == expected
    if operator == "not_equals":
        return row_value != expected
    if operator == "contains":
        if row_value is None:
            return False
        return str(expected).lower() in str(row_value).lower()
    if operator == "greater_than":
        return row_value is not None and row_value > expected
    if operator == "greater_or_equal":
        return row_value is not None and row_value >= expected
    if operator == "less_than":
        return row_value is not None and row_value < expected
    if operator == "less_or_equal":
        return row_value is not None and row_value <= expected
    if operator == "in":
        return row_value in expected if isinstance(expected, list) else False
    if operator == "not_in":
        return row_value not in expected if isinstance(expected, list) else True
    return True


def _apply_filters(rows: list[dict[str, Any]], definition: ReportDefinition) -> list[dict[str, Any]]:
    field_types = {field.id: field.type for field in DATASET_MAP[definition.dataset].fields}
    filtered = rows
    for rule in definition.filters:
        expected = _coerce_filter_value(field_types.get(rule.field, "string"), rule.value)
        filtered = [row for row in filtered if _match_filter(row.get(rule.field), rule.operator, expected)]
    return filtered


def _apply_sort(rows: list[dict[str, Any]], definition: ReportDefinition) -> list[dict[str, Any]]:
    sorted_rows = list(rows)
    for sort_rule in reversed(definition.sort):
        sorted_rows.sort(
            key=lambda row: (row.get(sort_rule.field) is None, row.get(sort_rule.field)),
            reverse=sort_rule.direction == "desc",
        )
    return sorted_rows


def _metric_value(rows: list[dict[str, Any]], metric: ReportMetricRequest) -> Any:
    if metric.operation == "count":
        return len(rows)
    if metric.operation == "count_distinct":
        return len({row.get(metric.field) for row in rows if row.get(metric.field) is not None})
    numeric_values = [Decimal(str(row.get(metric.field) or 0)) for row in rows if row.get(metric.field) is not None]
    if metric.operation == "sum":
        return sum(numeric_values, Decimal("0"))
    if metric.operation == "avg":
        return (sum(numeric_values, Decimal("0")) / Decimal(len(numeric_values))) if numeric_values else Decimal("0")
    return Decimal("0")


def _build_summary(rows: list[dict[str, Any]], definition: ReportDefinition) -> list[ReportMetricValue]:
    metric_requests = definition.metrics or [
        ReportMetricRequest(operation=metric.operation, field=metric.field, label=metric.label)
        for metric in DATASET_MAP[definition.dataset].metrics
    ]
    return [ReportMetricValue(label=_metric_label(metric), value=_fmt_value(_metric_value(rows, metric))) for metric in metric_requests]


def _build_grouped_rows(rows: list[dict[str, Any]], definition: ReportDefinition) -> tuple[list[str], list[dict[str, Any]]]:
    metrics = definition.metrics or [
        ReportMetricRequest(operation=metric.operation, field=metric.field, label=metric.label)
        for metric in DATASET_MAP[definition.dataset].metrics
    ]
    if not definition.group_by:
        grouped_row = {(_metric_label(metric)): _metric_value(rows, metric) for metric in metrics}
        columns = list(grouped_row.keys())
        return columns, [grouped_row]

    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[tuple(row.get(field) for field in definition.group_by)].append(row)

    result_rows: list[dict[str, Any]] = []
    columns = [*definition.group_by, *[_metric_label(metric) for metric in metrics]]
    for key, group_rows in grouped.items():
        item = {field: key[idx] for idx, field in enumerate(definition.group_by)}
        for metric in metrics:
            item[_metric_label(metric)] = _metric_value(group_rows, metric)
        result_rows.append(item)
    return columns, result_rows


def _project_rows(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    return [{column: row.get(column) for column in columns} for row in rows]


def _dedupe_rows_by_columns(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    unique_rows: list[dict[str, Any]] = []
    for row in rows:
        key = tuple(_serialize_value(row.get(column)) for column in columns)
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows



def _is_active_in_range(record: Any, date_from: date, date_to: date) -> bool:
    """True if the record's validity window overlaps the requested range."""
    valid_from: date | None = getattr(record, "valid_from", None)
    valid_to: date | None = getattr(record, "valid_to", None)
    # Starts before range ends AND ends after range starts (or open-ended)
    if valid_from and valid_from > date_to:
        return False
    if valid_to and valid_to <= date_from:
        return False
    return True


def _period_duration(valid_from: date | None, valid_to: date | None, date_from: date, date_to: date) -> int:
    """Days the record was active inside the requested range."""
    start = max(valid_from or date.min, date_from)
    end = min(valid_to or date.max, date_to)
    return max(0, (end - start).days)


async def _load_range_rows(db: AsyncSession, date_from: date, date_to: date, dataset_id: str) -> list[dict[str, Any]]:
    """
    Range mode: return one row per *version* of each entity that was active during
    [date_from, date_to].  Adds period_valid_from / period_valid_to / period_duration_days.
    Supported for tenant-centric datasets (tenant_snapshot_full, tenant_module_snapshot_full,
    master_dataset).  Falls back to as_of snapshot for other datasets.
    """
    if dataset_id not in {
        "tenant_snapshot_full", "tenant_module_snapshot_full", "master_dataset",
    }:
        # For non-temporal datasets just snapshot at date_to
        snapshot = await _load_snapshot_rows(db, date_to)
        rows = snapshot[dataset_id]
        # attach period fields
        for row in rows:
            row["period_valid_from"] = date_from
            row["period_valid_to"] = date_to
            row["period_duration_days"] = (date_to - date_from).days
        return rows

    # ── load all raw records ──────────────────────────────────────────────────
    tenant_result = await db.execute(select(Tenant))
    identity_result = await db.execute(select(TenantIdentity))
    contact_result = await db.execute(select(TenantContact).where(TenantContact.contact_type == "main"))
    address_result = await db.execute(select(TenantAddress).where(TenantAddress.addr_type == "main"))
    status_result = await db.execute(select(TenantStatus))
    subscription_result = await db.execute(select(TenantSubscription))
    subscription_module_result = await db.execute(select(TenantSubscriptionModule))
    module_result = await db.execute(select(Module))
    module_price_result = await db.execute(select(ModulePrice))
    template_result = await db.execute(select(OrgTemplate))
    template_default_result = await db.execute(select(OrgTemplateDefault))
    template_module_result = await db.execute(select(OrgTemplateModule))
    admin_user_result = await db.execute(select(AdminUser))

    tenants = {t.tenant_id: t for t in tenant_result.scalars().all()}
    modules = {m.slug: m for m in module_result.scalars().all()}
    templates = {t.id: t for t in template_result.scalars().all()}
    prices_all = module_price_result.scalars().all()
    template_defaults = template_default_result.scalars().all()
    template_modules = template_module_result.scalars().all()
    admin_users = {u.id: u for u in admin_user_result.scalars().all()}

    template_default_counts: dict[uuid.UUID, int] = defaultdict(int)
    for row in template_defaults:
        template_default_counts[row.template_id] += 1
    template_module_counts: dict[uuid.UUID, int] = defaultdict(int)
    for row in template_modules:
        template_module_counts[row.template_id] += 1

    # index helpers
    def _best_price(slug: str, as_of: date) -> Any:
        candidates = [p for p in prices_all if p.module_slug == slug and _effective(p, as_of)]
        if not candidates:
            return None
        return max(candidates, key=lambda p: getattr(p, "valid_from", date.min) or date.min)

    # ── collect all subscription versions active in range ─────────────────────
    subscriptions_in_range = [
        s for s in subscription_result.scalars().all()
        if _is_active_in_range(s, date_from, date_to)
    ]

    # All identities / contacts / addresses / statuses — keyed by (tenant_id, valid_from)
    all_identities = identity_result.scalars().all()
    all_contacts = contact_result.scalars().all()
    all_addresses = address_result.scalars().all()
    all_statuses = status_result.scalars().all()

    def _pick_for_date(records: list[Any], tenant_id: uuid.UUID, as_of: date) -> Any | None:
        candidates = [
            r for r in records
            if getattr(r, "tenant_id", None) == tenant_id and _effective(r, as_of)
        ]
        if not candidates:
            return None
        return max(candidates, key=lambda r: getattr(r, "valid_from", date.min) or date.min)

    # ── all subscription modules active in range ──────────────────────────────
    all_sub_modules = subscription_module_result.scalars().all()

    # derive (tenant_id -> [(sub, [active_modules_in_period])])
    result_rows: list[dict[str, Any]] = []

    if dataset_id in {"tenant_snapshot_full", "master_dataset"}:
        for sub in subscriptions_in_range:
            tenant = tenants.get(sub.tenant_id)
            if not tenant:
                continue
            # Use midpoint of overlap to pick contextual records
            overlap_start = max(sub.valid_from or date_from, date_from)
            overlap_end = min(sub.valid_to or date_to, date_to)
            mid = overlap_start + (overlap_end - overlap_start) / 2

            identity = _pick_for_date(all_identities, sub.tenant_id, mid)
            contact = _pick_for_date(all_contacts, sub.tenant_id, mid)
            address = _pick_for_date(all_addresses, sub.tenant_id, mid)
            status = _pick_for_date(all_statuses, sub.tenant_id, mid)
            template = templates.get(getattr(sub, "template_id", None))

            # Effective modules for this subscription in this period
            sub_modules = [
                m for m in all_sub_modules
                if m.tenant_subscription_id == sub.id and _is_active_in_range(m, overlap_start, overlap_end)
            ]
            derived_seat_count, derived_module_slugs = derive_subscription_snapshot(sub_modules)
            module_names = [
                modules[m.module_slug].name if m.module_slug in modules else m.module_slug
                for m in sub_modules
            ]

            tenant_name = getattr(identity, "name_he", None) or f"Tenant {tenant.org_number}"
            duration_days = _period_duration(sub.valid_from, sub.valid_to, date_from, date_to)

            row: dict[str, Any] = {
                "period_valid_from": overlap_start,
                "period_valid_to": overlap_end if sub.valid_to else None,
                "period_duration_days": duration_days,
                "record_type": "tenant_snapshot",
                "record_key": str(sub.tenant_id),
                "master_created_at": tenant.created_at,
                "tenant_org_number": tenant.org_number,
                "tenant_id": tenant.tenant_id,
                "org_number": tenant.org_number,
                "tenant_created_at": tenant.created_at,
                "identity_name_he": tenant_name,
                "identity_name_en": getattr(identity, "name_en", None),
                "identity_tax_id": getattr(identity, "tax_id", None),
                "identity_entity_type": getattr(identity, "entity_type", None),
                "identity_industry_code": getattr(identity, "industry_code", None),
                "identity_valid_from": getattr(identity, "valid_from", None),
                "identity_valid_to": getattr(identity, "valid_to", None),
                "contact_main_name": getattr(contact, "contact_name", None),
                "contact_main_email": getattr(contact, "email", None),
                "contact_main_phone": getattr(contact, "phone", None),
                "contact_main_website": getattr(contact, "website", None),
                "address_main_street": getattr(address, "street", None),
                "address_main_city": getattr(address, "city", None),
                "address_main_zip_code": getattr(address, "zip_code", None),
                "address_main_country": getattr(address, "country", None),
                "status_value": getattr(status, "status", None),
                "status_reason": getattr(status, "reason", None),
                "status_notes": getattr(status, "notes", None),
                "status_valid_from": getattr(status, "valid_from", None),
                "status_valid_to": getattr(status, "valid_to", None),
                "subscription_id": sub.id,
                "subscription_billing_cycle": getattr(sub, "billing_cycle", None),
                "subscription_currency": getattr(sub, "currency", None),
                "subscription_template_id": getattr(sub, "template_id", None),
                "subscription_template_name": getattr(template, "name", None),
                "subscription_seat_count": derived_seat_count,
                "subscription_selected_module_slugs": _join_text(derived_module_slugs),
                "subscription_discount_pct": float(getattr(sub, "discount_pct", 0) or 0),
                "subscription_is_price_locked": getattr(sub, "is_price_locked", False),
                "subscription_next_renewal_at": getattr(sub, "next_renewal_at", None),
                "subscription_valid_from": sub.valid_from,
                "subscription_valid_to": sub.valid_to,
                "module_count": len(sub_modules),
                "module_names": _join_text(module_names),
            }
            result_rows.append(row)

    elif dataset_id == "tenant_module_snapshot_full":
        for mod_row in all_sub_modules:
            if not _is_active_in_range(mod_row, date_from, date_to):
                continue
            # find parent subscription
            sub = next((s for s in subscriptions_in_range if s.id == mod_row.tenant_subscription_id), None)
            if sub is None:
                # subscription may have expired; still include it
                sub_all = [s for s in subscription_result.scalars().all() if s.id == mod_row.tenant_subscription_id]
                sub = sub_all[0] if sub_all else None
            if sub is None:
                continue
            tenant = tenants.get(sub.tenant_id)
            if not tenant:
                continue
            overlap_start = max(mod_row.valid_from or date_from, date_from)
            overlap_end = min(mod_row.valid_to or date_to, date_to)
            mid = overlap_start + (overlap_end - overlap_start) / 2
            identity = _pick_for_date(all_identities, sub.tenant_id, mid)
            status = _pick_for_date(all_statuses, sub.tenant_id, mid)
            module = modules.get(mod_row.module_slug)
            price = _best_price(mod_row.module_slug, mid)
            template = templates.get(getattr(sub, "template_id", None))
            tenant_name = getattr(identity, "name_he", None) or f"Tenant {tenant.org_number}"

            row = {
                "period_valid_from": overlap_start,
                "period_valid_to": overlap_end if mod_row.valid_to else None,
                "period_duration_days": _period_duration(mod_row.valid_from, mod_row.valid_to, date_from, date_to),
                "tenant_id": tenant.tenant_id,
                "org_number": tenant.org_number,
                "tenant_name": tenant_name,
                "tenant_status": getattr(status, "status", None),
                "subscription_id": sub.id,
                "subscription_template_name": getattr(template, "name", None),
                "subscription_billing_cycle": getattr(sub, "billing_cycle", None),
                "subscription_currency": getattr(sub, "currency", None),
                "subscription_seat_count": 0,
                "subscription_next_renewal_at": getattr(sub, "next_renewal_at", None),
                "module_assignment_id": mod_row.id,
                "tenant_subscription_id": mod_row.tenant_subscription_id,
                "module_slug": mod_row.module_slug,
                "module_name": getattr(module, "name", mod_row.module_slug),
                "module_description": getattr(module, "description", None),
                "module_is_required": getattr(module, "is_required", None),
                "module_is_active": getattr(module, "is_active", None),
                "module_sort_order": getattr(module, "sort_order", None),
                "source_type": mod_row.source_type,
                "module_status": mod_row.status,
                "module_seats": int(mod_row.seats or 0),
                "pricing_mode": mod_row.pricing_mode,
                "base_price_ils": float(getattr(price, "base_price_ils", 0) or 0),
                "per_seat_ils": float(getattr(price, "per_seat_ils", 0) or 0),
                "included_seats": int(getattr(price, "included_seats", 0) or 0),
                "setup_fee_ils": float(getattr(price, "setup_fee_ils", 0) or 0),
                "override_base_price_ils": float(getattr(mod_row, "override_base_price_ils", 0) or 0),
                "override_per_seat_ils": float(getattr(mod_row, "override_per_seat_ils", 0) or 0),
                "override_setup_fee_ils": float(getattr(mod_row, "override_setup_fee_ils", 0) or 0),
                "notes": getattr(mod_row, "notes", None),
                "valid_from": mod_row.valid_from,
                "valid_to": mod_row.valid_to,
                "next_renewal_at": getattr(sub, "next_renewal_at", None),
            }
            result_rows.append(row)

    # Sort by period_valid_from desc so newest first
    result_rows.sort(key=lambda r: (r.get("period_valid_from") or date.min), reverse=True)
    return result_rows


async def execute_report_query(db: AsyncSession, request: ReportQueryRequest) -> ReportResult:

    definition = _normalize_definition(request.definition)
    if definition.dataset not in DATASET_MAP:
        raise ValueError("Unknown dataset")
    if definition.limit < 1:
        raise ValueError("Limit must be positive")

    # ── Range mode: date_from + date_to both set ──────────────────────────────
    if definition.date_from and definition.date_to:
        source_rows = await _load_range_rows(db, definition.date_from, definition.date_to, definition.dataset)
    else:
        as_of = definition.as_of_date or date.today()
        source_rows = (await _load_snapshot_rows(db, as_of))[definition.dataset]

    filtered_rows = _apply_filters(source_rows, definition)
    filtered_rows = _apply_sort(filtered_rows, definition)

    if definition.view_mode == "summary":
        summary = _build_summary(filtered_rows, definition)
        columns, grouped_rows = _build_grouped_rows(filtered_rows, definition)
        grouped_rows = _apply_sort(grouped_rows, ReportDefinition(dataset=definition.dataset, sort=definition.sort))
        total = len(grouped_rows)
        paged_rows = grouped_rows[definition.offset:definition.offset + min(definition.limit, 1000)]
        return ReportResult(
            columns=columns,
            rows=[{key: _fmt_value(value) for key, value in row.items()} for row in paged_rows],
            total=total,
            summary=summary,
            applied_definition=definition,
        )

    columns = definition.columns or DATASET_MAP[definition.dataset].default_columns
    display_rows = _dedupe_rows_by_columns(filtered_rows, columns)
    summary = _build_summary(display_rows, definition)
    total = len(display_rows)
    paged_rows = _project_rows(display_rows[definition.offset:definition.offset + min(definition.limit, 1000)], columns)
    return ReportResult(
        columns=columns,
        rows=[{key: _fmt_value(value) for key, value in row.items()} for row in paged_rows],
        total=total,
        summary=summary,
        applied_definition=definition,
    )


def _render_csv(title: str | None, result: ReportResult) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([title or "CLICK Reports"])
    writer.writerow(["Generated at", datetime.now(UTC).isoformat()])
    writer.writerow(["Dataset", result.applied_definition.dataset])
    writer.writerow(["Total", result.total])
    writer.writerow([])
    if result.summary:
        writer.writerow(["Summary"])
        for metric in result.summary:
            writer.writerow([metric.label, metric.value])
        writer.writerow([])
    if result.columns:
        writer.writerow(result.columns)
    for row in result.rows:
        writer.writerow([row.get(column, "") for column in result.columns])
    return output.getvalue()


def _render_xlsx(title: str | None, result: ReportResult) -> bytes:
    """Generate a professional Excel (.xlsx) file with styled headers and columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "דוח")[:31]  # Excel sheet name limit
    ws.sheet_view.rightToLeft = True

    # Style definitions
    title_font = Font(name="Arial", bold=True, size=14, color="0F172A")
    header_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1E40AF")  # Deep blue
    summary_label_font = Font(name="Arial", bold=True, size=9, color="374151")
    summary_fill = PatternFill(fill_type="solid", fgColor="EFF6FF")
    alt_row_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    border_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=False, readingOrder=2)

    current_row = 1

    # Title row
    ws.cell(row=current_row, column=1, value=title or "CLICK Reports").font = title_font
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Meta rows
    ws.cell(row=current_row, column=1, value="הופק ב:").font = Font(name="Arial", size=8, color="64748B")
    ws.cell(row=current_row, column=2, value=datetime.now(UTC).strftime("%Y-%m-%d %H:%M"))
    current_row += 1
    ws.cell(row=current_row, column=1, value="מקור נתונים:").font = Font(name="Arial", size=8, color="64748B")
    ws.cell(row=current_row, column=2, value=result.applied_definition.dataset)
    current_row += 1
    ws.cell(row=current_row, column=1, value="סה\"כ שורות:").font = Font(name="Arial", size=8, color="64748B")
    ws.cell(row=current_row, column=2, value=result.total)
    current_row += 2  # blank row

    # Summary section
    if result.summary:
        ws.cell(row=current_row, column=1, value="סיכום").font = Font(name="Arial", bold=True, size=10, color="1E40AF")
        current_row += 1
        for metric in result.summary:
            lbl_cell = ws.cell(row=current_row, column=1, value=metric.label)
            lbl_cell.font = summary_label_font
            lbl_cell.fill = summary_fill
            lbl_cell.alignment = right_align
            val_cell = ws.cell(row=current_row, column=2, value=metric.value)
            val_cell.fill = summary_fill
            val_cell.alignment = right_align
            current_row += 1
        current_row += 1  # blank row

    # Header row
    if result.columns:
        for col_idx, col_name in enumerate(result.columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = center_align
        ws.row_dimensions[current_row].height = 20
        header_row = current_row
        current_row += 1

        # Data rows
        for row_idx, row in enumerate(result.rows):
            is_alt = row_idx % 2 == 1
            for col_idx, col_name in enumerate(result.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=row.get(col_name, ""))
                cell.border = cell_border
                cell.alignment = right_align
                if is_alt:
                    cell.fill = alt_row_fill
            current_row += 1

        # Auto-size columns (estimate)
        for col_idx, col_name in enumerate(result.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(col_name),
                *(len(str(row.get(col_name, ""))) for row in result.rows[:50]) if result.rows else [0],
            )
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Auto filter
        if result.rows:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(result.columns))}{header_row}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _render_html(title: str | None, result: ReportResult, dataset_label: str = "") -> str:
    """Generate a standalone, styled HTML report with RTL support."""
    report_title = html_module.escape(title or "CLICK Reports")
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M")

    # Build summary cards
    summary_html = ""
    if result.summary:
        cards = "".join(
            f'<div class="summary-card"><div class="summary-value">{html_module.escape(m.value)}</div>'
            f'<div class="summary-label">{html_module.escape(m.label)}</div></div>'
            for m in result.summary
        )
        summary_html = f'<div class="summary-grid">{cards}</div>'

    # Build table
    header_cells = "".join(f"<th>{html_module.escape(col)}</th>" for col in result.columns)
    row_html_parts = []
    for idx, row in enumerate(result.rows):
        cls = "alt" if idx % 2 == 1 else ""
        cells = "".join(
            f'<td>{html_module.escape(str(row.get(col, "—")))}</td>' for col in result.columns
        )
        row_html_parts.append(f'<tr class="{cls}"><td class="row-num">{idx + 1}</td>{cells}</tr>')
    rows_html = "\n".join(row_html_parts)

    num_col_header = '<th class="row-num">#</th>'

    return f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>{report_title}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    direction: rtl;
    text-align: right;
    background: #f8fafc;
    color: #0f172a;
    margin: 0;
    padding: 24px;
  }}
  .report-header {{
    background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%);
    color: white;
    padding: 24px 32px;
    border-radius: 12px;
    margin-bottom: 24px;
    box-shadow: 0 4px 16px rgba(30,64,175,0.2);
  }}
  .report-header h1 {{ margin: 0 0 4px; font-size: 22px; font-weight: 700; }}
  .report-meta {{ font-size: 12px; opacity: 0.8; display: flex; gap: 16px; flex-wrap: wrap; margin-top: 8px; }}
  .report-meta span {{ display: flex; align-items: center; gap: 4px; }}
  .summary-grid {{
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    margin-bottom: 24px;
  }}
  .summary-card {{
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 14px 20px;
    min-width: 140px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  }}
  .summary-value {{ font-size: 24px; font-weight: 700; color: #1e40af; }}
  .summary-label {{ font-size: 11px; color: #64748b; margin-top: 2px; text-transform: uppercase; letter-spacing: 0.05em; }}
  .table-wrapper {{
    background: white;
    border-radius: 12px;
    border: 1px solid #e2e8f0;
    overflow: auto;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }}
  thead {{
    position: sticky;
    top: 0;
    z-index: 2;
  }}
  thead tr {{
    background: #1e40af;
    color: white;
  }}
  th {{
    padding: 10px 14px;
    font-weight: 600;
    white-space: nowrap;
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }}
  td {{
    padding: 9px 14px;
    border-bottom: 1px solid #f1f5f9;
    color: #334155;
    white-space: nowrap;
  }}
  tr:hover td {{ background: #eff6ff; }}
  tr.alt td {{ background: #f8fafc; }}
  tr.alt:hover td {{ background: #eff6ff; }}
  .row-num {{
    color: #94a3b8;
    font-size: 11px;
    text-align: center;
    min-width: 36px;
  }}
  .report-footer {{
    margin-top: 20px;
    text-align: center;
    font-size: 11px;
    color: #94a3b8;
  }}
  @media print {{
    body {{ background: white; padding: 8px; }}
    .report-header {{ background: #1e40af !important; -webkit-print-color-adjust: exact; }}
  }}
</style>
</head>
<body>
<div class="report-header">
  <h1>{report_title}</h1>
  <div class="report-meta">
    <span>&#128197; הופק ב: {generated_at}</span>
    <span>&#128202; מקור: {html_module.escape(dataset_label or result.applied_definition.dataset)}</span>
    <span>&#128203; סה&quot;כ: {result.total:,} שורות</span>
  </div>
</div>
{summary_html}
<div class="table-wrapper">
  <table>
    <thead><tr>{num_col_header}{header_cells}</tr></thead>
    <tbody>
{rows_html}
    </tbody>
  </table>
</div>
<div class="report-footer">CLICK HR &mdash; מחולל דוחות &mdash; {generated_at}</div>
</body>
</html>"""


def _render_pdf(title: str | None, result: ReportResult) -> bytes:
    _register_font()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportBody", parent=styles["BodyText"], fontName=FONT_NAME, fontSize=9, leading=12, alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Heading1"], fontName=FONT_NAME, fontSize=18, leading=22, alignment=TA_RIGHT, textColor=colors.HexColor("#0f172a")))
    story = [
        Paragraph(_rtl(title or "CLICK Reports"), styles["ReportTitle"]),
        Spacer(1, 4),
        Paragraph(_rtl(f"Dataset: {result.applied_definition.dataset}"), styles["ReportBody"]),
        Spacer(1, 8),
    ]
    if result.summary:
        summary_table = Table(
            [[Paragraph(_rtl(metric.label), styles["ReportBody"]), Paragraph(_rtl(metric.value), styles["ReportBody"])] for metric in result.summary],
            colWidths=[55 * mm, 115 * mm],
            hAlign="RIGHT",
        )
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ]))
        story.extend([summary_table, Spacer(1, 10)])
    if result.rows:
        data = [[Paragraph(_rtl(column), styles["ReportBody"]) for column in result.columns]]
        for row in result.rows:
            data.append([Paragraph(_rtl(str(row.get(column, "—"))), styles["ReportBody"]) for column in result.columns])
        width = 180 * mm / max(len(result.columns), 1)
        table = Table(data, colWidths=[width] * len(result.columns), repeatRows=1, hAlign="RIGHT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ]))
        story.append(table)
    doc.build(story)
    return buffer.getvalue()


async def export_report(db: AsyncSession, request: ReportExportRequest) -> ReportExportResponse:
    result = await execute_report_query(db, ReportQueryRequest(title=request.title, definition=request.definition))
    safe_name = (request.title or request.definition.dataset).lower().replace(" ", "-").replace("/", "-")
    # For xlsx and html we need all rows — re-run with high limit
    if request.format in ("xlsx", "html"):
        export_def = request.definition.model_copy(update={"limit": 5000, "offset": 0})
        result = await execute_report_query(db, ReportQueryRequest(title=request.title, definition=export_def))
    dataset_label = DATASET_MAP.get(request.definition.dataset, None)
    dataset_label_str = dataset_label.label if dataset_label else request.definition.dataset
    if request.format == "csv":
        raw = _render_csv(request.title, result)
        return ReportExportResponse(
            file_name=f"{safe_name}.csv",
            mime_type="text/csv; charset=utf-8",
            content_base64=base64.b64encode(raw.encode("utf-8-sig")).decode("ascii"),
        )
    if request.format == "xlsx":
        raw_bytes = _render_xlsx(request.title, result)
        return ReportExportResponse(
            file_name=f"{safe_name}.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content_base64=base64.b64encode(raw_bytes).decode("ascii"),
        )
    if request.format == "html":
        raw_html = _render_html(request.title, result, dataset_label_str)
        return ReportExportResponse(
            file_name=f"{safe_name}.html",
            mime_type="text/html; charset=utf-8",
            content_base64=base64.b64encode(raw_html.encode("utf-8")).decode("ascii"),
        )
    raw = _render_pdf(request.title, result)
    return ReportExportResponse(
        file_name=f"{safe_name}.pdf",
        mime_type="application/pdf",
        content_base64=base64.b64encode(raw).decode("ascii"),
    )


def _saved_to_out(row: SavedReportView, owner_name: str | None) -> SavedReportViewOut:
    definition = _normalize_definition(ReportDefinition.model_validate(row.definition_json))
    return SavedReportViewOut(
        id=row.id,
        name=row.name,
        description=row.description,
        kind=getattr(row, "kind", "report"),
        dataset=definition.dataset,
        visibility=row.visibility,
        owner_id=row.owner_id,
        owner_name=owner_name,
        created_at=row.created_at,
        updated_at=row.updated_at,
        definition=definition,
    )


async def list_saved_reports(db: AsyncSession, current_user: CurrentUser, kind: str | None = None) -> list[SavedReportViewOut]:
    query = (
        select(SavedReportView, AdminUser.full_name)
        .join(AdminUser, AdminUser.id == SavedReportView.owner_id)
        .where((SavedReportView.owner_id == current_user.id) | (SavedReportView.visibility == "shared"))
        .order_by(SavedReportView.updated_at.desc().nullslast(), SavedReportView.created_at.desc())
    )
    if kind in {"report", "template"}:
        query = query.where(SavedReportView.kind == kind)
    try:
        result = await db.execute(query)
    except ProgrammingError as exc:
        if "saved_report_views" in str(exc).lower():
            return []
        raise
    rows = [_saved_to_out(row, owner_name) for row, owner_name in result.all()]
    if kind in {"report", "template"}:
        rows = [row for row in rows if row.kind == kind]
    return rows


async def create_saved_report(db: AsyncSession, body: SavedReportViewCreate, current_user: CurrentUser) -> SavedReportViewOut:
    definition = _normalize_definition(body.definition)
    if definition.dataset not in DATASET_MAP:
        raise ValueError("Unknown dataset")
    row = SavedReportView(
        name=body.name,
        description=body.description,
        dataset=definition.dataset,
        definition_json=definition.model_dump(mode="json"),
        kind=body.kind,
        visibility=body.visibility,
        owner_id=current_user.id,
    )
    db.add(row)
    await db.flush()
    await db.refresh(row)
    result = await db.execute(select(AdminUser.full_name).where(AdminUser.id == current_user.id))
    return _saved_to_out(row, result.scalar_one_or_none())


async def update_saved_report(
    db: AsyncSession,
    report_id: uuid.UUID,
    body: SavedReportViewUpdate,
    current_user: CurrentUser,
) -> SavedReportViewOut:
    result = await db.execute(select(SavedReportView).where(SavedReportView.id == report_id))
    row = result.scalar_one_or_none()
    if row is None:
        raise ValueError("Saved report not found")
    if row.owner_id != current_user.id:
        raise PermissionError("Only the owner can edit this saved report")
    if body.name is not None:
        row.name = body.name
    if body.description is not None:
        row.description = body.description
    if body.kind is not None:
        row.kind = body.kind
    if body.visibility is not None:
        row.visibility = body.visibility
    if body.definition is not None:
        definition = _normalize_definition(body.definition)
        if definition.dataset not in DATASET_MAP:
            raise ValueError("Unknown dataset")
        row.dataset = definition.dataset
        row.definition_json = definition.model_dump(mode="json")
    row.updated_at = datetime.now(UTC)
    await db.flush()
    await db.refresh(row)
    owner_name_result = await db.execute(select(AdminUser.full_name).where(AdminUser.id == row.owner_id))
    return _saved_to_out(row, owner_name_result.scalar_one_or_none())


async def run_saved_report(db: AsyncSession, report_id: uuid.UUID, current_user: CurrentUser) -> ReportResult:
    result = await db.execute(select(SavedReportView).where(SavedReportView.id == report_id))
    row = result.scalar_one_or_none()
    if row is None:
        raise ValueError("Saved report not found")
    if row.visibility != "shared" and row.owner_id != current_user.id:
        raise PermissionError("Saved report is not available")
    if getattr(row, "kind", "report") != "report":
        raise ValueError("Saved template cannot be run directly")
    definition = _normalize_definition(ReportDefinition.model_validate(row.definition_json))
    return await execute_report_query(db, ReportQueryRequest(title=row.name, definition=definition))


async def delete_saved_report(db: AsyncSession, report_id: uuid.UUID, current_user: CurrentUser) -> None:
    result = await db.execute(select(SavedReportView).where(SavedReportView.id == report_id))
    row = result.scalar_one_or_none()
    if row is None:
        raise ValueError("Saved report not found")
    if row.owner_id != current_user.id:
        raise PermissionError("Only the owner can delete this saved item")
    await db.delete(row)
